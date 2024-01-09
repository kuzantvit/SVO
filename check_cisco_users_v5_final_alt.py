import paramiko
import getpass
import sys
import time
import re
from AD import ADinfo
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

import datetime


now = datetime.datetime.now()
if now.month < 10:
        month_string = '0' + str(now.month)
else:
        month_string = str(now.month)

if now.day < 10:
        day_string = '0' + str(now.day)
else:
        day_string = str(now.day)
date_string = day_string + month_string

result_list = open('c:\\Users\\avkuzmin\\Documents\\scripts\\network\\workingcsvresult'+ date_string + '_' + str(now.hour) + '-' + str(now.minute) + '.txt', mode ='a', encoding = 'utf-8')



#COMMAND = sys.argv[1]
COMMAND = 'show run | i username'
command2 = 'show arp'
#USER = input('Username: ')
USER = 'avkuzmin'
#PASSWORD = getpass.getpass()
PASSWORD = '5353535353535353535353535353535'
#ENABLE_PASS = getpass.getpass(prompt='Enter enable password: ')
#ENABLE_PASS =  'On'
DEVICES_IP = ['8.8.8.8']
IP = '8.8.8.8'

#ip_tocheck = ['8.8.8.8']
#ip_tocheck = ['8.8.8.8']
ip_tocheck = []
#ip_tocheck = ['8.8.8.8', '8.8.8.8']
#ip_to_input = input('Vvedite ip address: ')
#ip_tocheck.append(ip_to_input)

wb_read = load_workbook(filename='c:\\Users\\avkuzmin\\Documents\\scripts\\network\\files\\network_devices.xlsx')
ws_read = wb_read['Лист1']
for i in range(600):
    if ws_read.cell(row=i+1,column=1).value != None:
        ip_tocheck.append(ws_read.cell(row=i+1,column=1).value)
fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
alignment_cell = Alignment(horizontal='left', vertical ='center', wrapText='True')





#client = paramiko.SSHClient()
#client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

#client.connect(hostname=IP, username=USER, password=PASSWORD,
#                   look_for_keys=False, allow_agent=False)
#ssh = client.invoke_shell(term='vt100', width=200, height=84, width_pixels=0, height_pixels=0, environment=None)



users_final = []
ad_analog_final = []
suspicious_users_final =[]
resik = []

result_list = open('c:\\Users\\avkuzmin\\Documents\\scripts\\network\\workingcsvresult'+ date_string + '_' + str(now.hour) + '-' + str(now.minute) + '.txt', mode ='a', encoding = 'utf-8')
def check_cisco_users(ip_tocheck):
    global IP
    global USER
    global PASSWORD
    global COMMAND
    print(ip_tocheck)
    result_array =[]
    users = []
    ad_analog_css =[]
    users_css =[]    
    ad_analog = []
    all_users = ['ciscolms', 'npetrov', 'nick', 'test']
    suspicious_users_list = []
    suspicious_users_css =[]
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())

    client.connect(hostname=IP, username=USER, password=PASSWORD,
                   look_for_keys=False, allow_agent=False)
    ssh = client.invoke_shell(term='vt100', width=240, height=120, width_pixels=0, height_pixels=0, environment=None)
    time.sleep(1)
    ssh.send('ssh 8.8.8.8\n')
    time.sleep(1)
    ssh.send('test\n')
    time.sleep(1)
    ssh.send('enable\n')
    time.sleep(1)
    ssh.send('test\n')
    time.sleep(1)
# 
    ssh.send('ssh -vrf MGMT ' + str(ip_tocheck) + '\n')
    time.sleep(8)
    ssh_out1 = ssh.recv(8000).decode('utf-8')
    check1 = ssh_out1[-10:]
    print(check1)
    if 'Password:' not in check1:
        users_css.append('pass issue with IP: ' + str(ip_tocheck) + ' !')
        ad_analog_css.append('pass issue IP: ' + str(ip_tocheck) + ' !')
        ad_analog.append('pass issue IP: ' + str(ip_tocheck) + ' !')
        suspicious_users_css.append('Issue')
        print('big issue')
        print('======================================Next itteration=============================================')
        return users_css, ad_analog_css, suspicious_users_css
#
    ssh.send('test\n')
    time.sleep(8)
    ssh_out2 = ssh.recv(8000).decode('utf-8')
    check2 = ssh_out2[-10:]
    print(check2)
    result_list.write(ssh_out2)
    if 'Password:' in check2:
        #
        ssh.send(PASSWORD + '\n')
        time.sleep(6)
        print('Local Pass')
#       
    ssh.send('enable\n')
    time.sleep(5)
    ssh_out_en = ssh.recv(8000).decode('utf-8')
    check_en = ssh_out_en[-10:]
    print(check_en)
    result_list.write(ssh_out_en)
    if 'Password:' not in check_en:
        #ssh.send(PASSWORD + '\n')
        #time.sleep(6)
        ssh.send(COMMAND + '\n')
        time.sleep(12)
        ssh.send('\n')
        ssh_out_if = ssh.recv(8000).decode('utf-8')
        users_css.append('issue with enable check manually: ' + str(ip_tocheck) + ' !')
        ad_analog_css.append('issue with enable check manually: ' + str(ip_tocheck) + ' !')
        ad_analog.append('issue with enable check manually: ' + str(ip_tocheck) + ' !')
        suspicious_users_css.append('Issue')
        print('issue with enable')
        result_list.write('issue with enable')
        result_list.write(ssh_out_if)
        print(ssh_out_if)
        print('======================================Next itteration=============================================')
        return users_css, ad_analog_css, suspicious_users_css
    #
    ssh.send('test\n')
    time.sleep(3)
    ssh_out3 = ssh.recv(8000).decode('utf-8')
    check3 = ssh_out3[-10:]
    print(check3)
    result_list.write(ssh_out3)
    if 'Error in' in ssh_out3 or 'Access denied' in ssh_out3:
        ssh.send('enable\n')
        time.sleep(5)
        ssh.send(PASSWORD + '\n')
        time.sleep(5)
        print('Local Pass for enable')
    ssh.send('\n')
    ssh.recv(8000).decode('utf-8')
    ssh.send(COMMAND + '\n')
    time.sleep(12)
    ssh.send('\n')
    result = ssh.recv(8000).decode('utf-8')
    result_list.write(result)
    print(result)
    if 'svoD_C6807-1#' in result:
        users_css.append('Issue with IP: ' + str(ip_tocheck) + ' !')
        ad_analog_css.append('Issue with IP: ' + str(ip_tocheck) + ' !')
        suspicious_users_css.append('Issue')
    elif 'closed by foreign host' in result:    
        users_css.append('Issue with IP: ' + str(ip_tocheck) + ' !')
        ad_analog_css.append('Issue with IP: ' + str(ip_tocheck) + ' !')
        suspicious_users_css.append('Issue')
    elif 'input detected' in result:
        users_css.append('Issue with IP (no access): ' + str(ip_tocheck) + ' !')
        ad_analog_css.append('Issue with IP(no access): ' + str(ip_tocheck) + ' !')
        suspicious_users_css.append('Issue')
    else:
            #result_list.write(result)
            result_array.append(result)
            ds = result.split('username ')
            for i in ds:
                if 'show run | i username' in i:
                    ds.pop(ds.index(i))
            for i in ds:
                result_re = re.search(r'^\S*', i)
                member = result_re.group()
                users.append(member)
                if member not in all_users:
                    suspicious_users_list.append(str(member))
            #if len(users) > 26:
            #    suspicious_users.append("YES")
            for user in users:
                username, surname = ADinfo.get_users_with_specified_attribute('sAMAccountName', user)
                ad_analog.append(surname[0])
            ad_analog_convert = ' \n'.join(ad_analog)
            users_css_convert = ' \n '.join(users)
            suspicious_users_convert = ' \n'.join(suspicious_users_list)
            ad_analog_css.append(ad_analog_convert)
            users_css.append(users_css_convert)
            suspicious_users_css.append(suspicious_users_convert)
    print('======================================Next itteration=============================================')
    if len(suspicious_users_css) == 0:
        suspicious_users_css.append('None')
    client.close()
    return users_css, ad_analog_css, suspicious_users_css
#ADinfo.save_data_to_excel(ip_tocheck, users, ad_analog, filename = 'c:\\Users\\avkuzmin\\Documents\\scripts\\network\\result\\test_users.xlsx')
for i in ip_tocheck:
    users_fin, ad_analog_fin, suspicious_users = check_cisco_users(i)
    users_final.append(users_fin[0])
    ad_analog_final.append(ad_analog_fin[0])
    suspicious_users_final.append(suspicious_users[0])
    #resik.append(result)

    
def save_data_to_excel(*args, filename='c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\results\\Проверка12.xlsx', **kwargs):
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(str(date_string))
    for counter, k in enumerate(kwargs):
        #print(counter)
        #print(k)
        #print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]
            ws_write.cell(row=counter+2, column=count_data+1).alignment = alignment_cell
            ws_write.cell(row=counter+2, column=count_data+1).border = thin_border
            if 'poroikov_dv' in (data[counter]):
                ws_write.cell(row=counter+2, column=count_data+1).fill = fill_cell_orange
                ws_write.cell(row=counter+2, column=5).value = 'poroikov_dv'

    wb_write.save(filename)
    return wb_write

    
save_data_to_excel(ip_tocheck, users_final, ad_analog_final, suspicious_users_final, filename = 'c:\\cisco_test\\cisco_users.xlsx', ip_address = 'ИП Адрес', local_users = 'Локальные УЗ', ad_users = ' Похожие УЗ из AD', suspicious_users = 'Подозрительные УЗ')
#ADinfo.save_data_to_excel(ip_tocheck, users_css, ad_analog_css, filename = 'c:\\Users\\av.kuzmin\\Documents\\scripts\\network\\result\\test_users.xlsx')
result_list.close()

wb_cisco_users_monday = load_workbook(filename='c:\\cisco_test\\cisco_users.xlsx')
ws_cisco_users_monday = wb_cisco_users_monday[str(date_string)]
wb_net_splunk = load_workbook(filename='C:\\cisco_test\\net_dev_splunk.xlsx')
ws_net_splunk = wb_net_splunk['Лист1']
for i in range(1000):
    for m in range(1000):
        if ws_cisco_users_monday.cell(row=m+1,column=1).value == ws_net_splunk.cell(row=i+1, column=3).value:
            #ws_IT.cell(row=i+3, column=44).value = ws_active.cell(row=m+4,column=4).value
            #ws_IT.cell(row=i+3, column=45).value = ws_active.cell(row=m+4,column=5).value
            ws_cisco_users_monday.cell(row=m+1,column=7).value =  ws_net_splunk.cell(row=i+1, column=1).value
            #print('Yes')
#ws_IT.cell(row=1, column=47).value = 
wb_cisco_users_monday.save(filename='c:\\cisco_test\\cisco_users_with_names.xlsx')



