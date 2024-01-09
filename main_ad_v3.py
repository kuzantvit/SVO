import simple_functions
from openpyxl import Workbook, load_workbook
from ADinfo import username_list, surname_list, getinfo_by_username, \
                   get_all_workers, getinfo_by_surname, get_status_by_username, \
                   get_status_by_surname, get_users_from_group
import datetime
now = datetime.datetime.now()
if now.month < 10:
        month_string = '0' + str(now.month)
else:
        month_string = str(now.month)

if now.day < 10:
        day_string = '0' + str(now.day)
else:
        day_string = str(now.day)
date_string = day_string + month_string


filename_users = 'c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\files\\export_usernames.csv'
filename_surnames = 'c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\files\\export_surnames.csv'
filename_all = 'c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\files\\Работники.xlsx'
result_wb = 'c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\results\\Проверка.xlsx'
wb = load_workbook(filename=result_wb)
ws = wb.create_sheet(date_string + '_' + str(now.hour) + '-' + str(now.minute))

info = []
status =[]
locked_users =[]
unlocked_users =[]
all_workers = get_all_workers(filename=filename_all)
work_in_mash =[]
department =[]
fired =[]
bad_users =[]
bad_users_cn =[]
company = []

check_method = input('How to check man/file: ')
if check_method == 'file':
    list_type = input("username/surname: ")
    if list_type == 'username':
        suspicious_users = username_list(filename=filename_users)
        for user in suspicious_users:
            info.append(getinfo_by_username(user))
            status.append(get_status_by_username(user))
            if isinstance(getinfo_by_username(user), str):
                bad_users.append(user)
            else:
                try:
                    department.append(getinfo_by_username(user).department.value)
                except:
                    department.append("NO department/service account")
                try:
                    company.append(getinfo_by_username(user).company.value)
                except:
                    company.append("NO company/service account")
                try:
                    check = getinfo_by_username(user).description.value
                except:
                    check = 'nothing'
                if get_status_by_username(user) == 'Locked':
                    locked_users.append(user)
                elif get_status_by_username(user) == 'Unlocked':
                    unlocked_users.append(user)
                if getinfo_by_username(user).name.value in all_workers:
                    work_in_mash.append(getinfo_by_username(user).name.value)
                if get_status_by_username(user) == 'Unlocked' and getinfo_by_username(user).name.value not in all_workers:
                    bad_users.append(getinfo_by_username(user).name.value)
                    bad_users_cn.append(user)
                if (getinfo_by_username(user).name.value in all_workers) or (check in all_workers):
                    pass
                else:
                    fired.append(getinfo_by_username(user).name.value)
    else:
        suspicious_users = surname_list(filename=filename_surnames)
        for user in suspicious_users:
            info.append(getinfo_by_surname(user))
            status.append(get_status_by_surname(user))
            if isinstance(getinfo_by_surname(user), str):
                bad_users.append(user)
            else:
                try:
                    department.append(getinfo_by_surname(user).department.value)
                except:
                    department.append("NO department/service account")
                try:
                    company.append(getinfo_by_surname(user).company.value)
                except:
                    company.append("NO company/service account")
                try:
                    check = getinfo_by_surname(user).description.value
                except:
                    check = 'nothing'
                if get_status_by_surname(user) == 'Locked':
                    locked_users.append(user)
                elif get_status_by_surname(user) == 'Unlocked':
                    unlocked_users.append(user)
                if getinfo_by_surname(user).name.value in all_workers:
                    work_in_mash.append(user)
                if get_status_by_surname(user) == 'Unlocked' and getinfo_by_surname(user).name.value not in all_workers:
                    bad_users.append(user)
                    bad_users_cn.append(getinfo_by_surname(user).sAMAccountName.value)
                if (getinfo_by_surname(user).name.value in all_workers) or (check in all_workers):
                    pass
                else:
                    fired.append(user)
else:
    list_type = input("username/surname: ")
    suspicious_users = []
    if list_type == 'username':
        user_input = input('Enter usernames/ press "q" to stop: ')
        while user_input != 'q':
            suspicious_users.append(user_input)
            user_input = input('Enter usernames/ press "q" to stop: ')
        for user in suspicious_users:
            info.append(getinfo_by_username(user))
            status.append(get_status_by_username(user))
            if isinstance(getinfo_by_username(user), str):
                bad_users.append(user)
            else:
                try:
                    department.append(getinfo_by_username(user).department.value)
                except:
                    department.append("Please check service/admin account")
                try:
                    company.append(getinfo_by_username(user).company.value)
                except:
                    company.append("Please check service/admin account")
                try:
                    check = getinfo_by_username(user).description.value
                except:
                    check = 'nothing'
                if get_status_by_username(user) == 'Locked':
                    locked_users.append(user)
                elif get_status_by_username(user) == 'Unlocked':
                    unlocked_users.append(user)
                if getinfo_by_username(user).name.value in all_workers:
                    work_in_mash.append(getinfo_by_username(user).name.value)
                if get_status_by_username(user) == 'Unlocked' and getinfo_by_username(user).name.value not in all_workers:
                    bad_users.append(getinfo_by_username(user).name.value)
                    bad_users_cn.append(user)
                if (getinfo_by_username(user).name.value in all_workers) or (check in all_workers):
                    pass
                else:
                    fired.append(getinfo_by_username(user).name.value)
    elif list_type == 'surname':
        user_input = input('Enter surrnames/ press "q" to stop: ')
        while user_input != 'q':
            suspicious_users.append(user_input)
            user_input = input('Enter surnames/ press "q" to stop: ')
        for user in suspicious_users:
            info.append(getinfo_by_surname(user))
            status.append(get_status_by_surname(user))
            if isinstance(getinfo_by_surname(user), str):
                bad_users.append(user)
            else:
                try:
                    department.append(getinfo_by_surname(user).department.value)
                except:
                    department.append("Please check service/admin account")
                try:
                    company.append(getinfo_by_surname(user).company.value)
                except:
                    company.append("Please check service/admin account")
                try:
                    check = getinfo_by_surname(user).description.value
                except:
                    check = 'nothing'
                if get_status_by_surname(user) == 'Locked':
                    locked_users.append(user)
                elif get_status_by_surname(user) == 'Unlocked':
                    unlocked_users.append(user)
                if getinfo_by_surname(user).name.value in all_workers:
                    work_in_mash.append(user)
                if get_status_by_surname(user) == 'Unlocked' and getinfo_by_surname(user).name.value not in all_workers:
                    bad_users.append(user)
                    bad_users_cn.append(getinfo_by_surname(user).sAMAccountName.value)
                if (getinfo_by_surname(user).name.value in all_workers) or (check in all_workers):
                    pass
                else:
                    fired.append(user)

groupname = input('Введите имя группы (q to exit): ')
user_from_group_department =[]
user_from_group_company = []
users_in_group =[]
if groupname =='q':
    pass
else:
    users_in_group = get_users_from_group(groupname)
    for user in users_in_group:
        try:
            user_from_group_department.append(getinfo_by_surname(user).department.value)
        except:
            user_from_group_department.append('Департамент отсутствует')


result_list = open('c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\results\\workingcsvresult'+ date_string + '_' + str(now.hour) + '-' + str(now.minute) + '.txt', mode ='a', encoding = 'utf-8')
result_list.write('\n=====================================================ЗАБЛОКИРОВАНЫЕ УЗ ИЗ СПИСКА=======================================================\n' + '\n'.join(locked_users))
result_list.write('\n=====================================================РАЗБЛОКИРОВАННЫЕ УЗ ИЗ СПИСКА=======================================================\n' + '\n'.join(unlocked_users))
result_list.write('\n=====================================================УВОЛЕННЫЕ СОТРУДНИКИ ИЗ СПИСКА=======================================================\n' + '\n'.join(fired))
result_list.write('\n=====================================================ПОДОЗРИТЕЛЬНЫЕ УЗ ИЗ СПИСКА=======================================================\n' + '\n'.join(bad_users))
result_list.write('\n=====================================================СПИСОК УЗ И ДЕПАРТАМЕНТЫ=======================================================\n')
for i in range(len(suspicious_users)):
    result_list.write(suspicious_users[i] + '\t' + department[i] + '\t\t\t\t\t' + company[i] + '\n')
result_list.write('\n=====================================================СПИСОК УЗ ИЗ ГРУППЫ '+ str(groupname) +' И ДЕПАРТАМЕНТЫ=======================================================\n')
for i in range(len(users_in_group)):
    result_list.write(users_in_group[i] + '\t\t\t' + user_from_group_department[i] + '\n')
result_list.close()
