import os
import os.path
import time
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side


now = datetime.datetime.now()
if now.month < 10:
        month_string = '0' + str(now.month)
else:
        month_string = str(now.month)

if now.day < 10:
        day_string = '0' + str(now.day)
else:
        day_string = str(now.day)
date_string = day_string + month_string
# это дата


result_list = open('C:\\cisco_test\\temp\\Result '+ 'net check ' + 'files_copied.txt', mode ='a')

def save_data_to_excel(*args, filename='c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\results\\Проверка12.xlsx', **kwargs):
    """ Функция для сохранения данных в excel
        """
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(str(date_string))
    for counter, k in enumerate(kwargs):
        #print(counter)
        #print(k)
        #print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]

    wb_write.save(filename)
    return wb_write

a =[]
b =[]
c =[]
for dirpath, dirnames, filename in os.walk('\\\\SOLOVYEVA\\c$\\Users\\oasoloveva\\Documents'):
        #print('1==========')
        print(dirpath)
        print('2==========')
        #print(dirnames)
        #print('3==========')
        #print(filename)
        #print('4==========')
        #a = '\n'.join(dirpath, dirnames, filename)
        a.append(dirpath)
        b.append(filename)
        for i in filename:
            c.append(os.path.join(dirpath, i))


g = '\n '.join(c)
var_1 = a
var_2 =[]
for m in b:
    var_2.append('\n '.join(m))

#result_list.write(g)
save_data_to_excel(var_1,var_2, filename = 'c:\\cisco_test\\solov_copied.xlsx')
result_list.close()
print(g)

"""
turn =[]
incr_dir = []
file_path = 'c:\\cisco_test\\'
with os.scandir(file_path) as listOfEntries:

    for entry in listOfEntries:
        if entry.is_dir():
            turn.append(file_path + '\\' + entry.name)
if entry.is_dir():
    incr_dir.append(entry)
if len(incr_dir) == 0:
    turn.append(file_path + '     ! no subfolders')
"""
