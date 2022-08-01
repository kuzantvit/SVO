from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl
import win32net, win32con
import os
import time
import subprocess
from subprocess import check_output
from subprocess import Popen, PIPE

def create_ip_list(scope_address, mask):
    A_list = scope_address.split('.')
    a4 =[x for x in range(256)]
    a2 = A_list[1]
    a1 = A_list[0]
    a3 = []
    a3.append(A_list[2])
    ip =[]
    mask = str(mask)
    if mask == '255.255.255.0':
        a3 = a3
    elif mask == '255.255.254.0':
        a3.append(str(int(A_list[2])+1))
    elif mask == '255.255.192.0':
        a3 = [x for x in range(int(A_list[2]), (int(A_list[2])+64))]
    elif mask =='255.255.252.0':
        a3 = [x for x in range(int(A_list[2]), (int(A_list[2])+4))]
    elif mask =='255.255.248.0':
        a3 = [x for x in range(int(A_list[2]), (int(A_list[2])+8))]
    elif mask =='255.255.224.0':
        a3 = [x for x in range(int(A_list[2]), (int(A_list[2])+32))]
    elif mask =='255.255.240.000':
        a3 = [x for x in range(int(A_list[2]), (int(A_list[2])+16))]
    elif mask =='255.255.128.000':
        a3 = [x for x in range(int(A_list[2]), (int(A_list[2])+128))]       
    elif mask =='test':
        a3 = a3
        a4=[x for x in range(15)]
    for i in a3:
        for m in a4:
            ip_name = a1 +'.'+ a2+ '.'+ str(i) + '.' + str(m)
            ip.append(ip_name)
    #print(ip)
    return ip

def host_folders(host):
    print("Checking following IP address:  " + str(host) + '\n')
    result ={}
    permission ={}
    hostname = ('\\\\'+str(host))
    turn =[]
    try:
        icheck=host.split()
    except:
        print('This is a date not a hostname')
    folder_list =[]
    permissions = []
    try:
        shares, total, what = win32net.NetShareEnum(hostname, 1)

        for i in range(len(shares)):
            a = shares[i].get('netname')
            if a not in ['ADMIN$','C$','D$','IPC$','print$', 'e$']:
                folder_list.append(a.lower())
            result.update({host:folder_list})

        for i in range(len(shares)):
            a = shares[i].get('netname')
            b = shares[i].get('type')
            if a not in ['ADMIN$','C$','D$','IPC$','print$', 'prnproc$'] and b==0:
                file_path = '\\\\' + host +'\\' + a
                try:
                    p = subprocess.run(['icacls', file_path], check = True, stdout = PIPE)
                    perm = p.stdout.decode('cp866', 'ignore')
                    permissions.append(perm)
                
                except:
                    permissions.append(file_path + ' - folder permission issue; no access or system folder')                
        if len(permissions) == 0:
            permissions.append('no suspicious folders')
        
        for i in range(len(shares)):
            a = shares[i].get('netname')
            b = shares[i].get('type')
            #print(shares)
            if a not in ['ADMIN$','C$','D$','IPC$','print$', 'prnproc$'] and b==0:
                file_path = '\\\\' + str(host) +'\\' + a.lower()
                try:
                    with os.scandir(file_path) as listOfEntries:
                        incr_dir = []
                        for entry in listOfEntries:
                            if entry.is_dir():
                                turn.append(file_path + '\\' + entry.name)
                            if entry.is_dir():
                                incr_dir.append(entry)
                        if len(incr_dir) == 0:
                            turn.append(file_path + '     ! no subfolders')
                except:
                    turn.append(file_path + ' - is not a file share or accessing subfolder issue \n')              
    except:
        result.update({host:'connection or access issue'})
        permissions.append('folder permission issue')
        turn.append('issue getting subfolders')
        print('Issue with following IP address:   ' + str(host) +'\n')
        
    return result, permissions, turn

def check_pc_locusers(host):
    #print("Checking following IP address:  " + str(host) + '\n')
    account_status = {}
    Account_check_result = {}
    user_groups = {}
    users = []
    membership = []
    printer=[]
    print('test')
    try:
        (user_listt,totall,ress)=win32net.NetGroupEnum(host,0,0)
        if len(user_listt) >= 2:
            groupName = user_listt[1]['name']    
        else:
            groupName = user_listt[0]['name']
        a,b,c = win32net.NetGroupGetUsers(host,groupName,1)
        for usrs in a:
            account_one_status = 'empty'
            username = usrs.get('name')
            
            group_member = win32net.NetUserGetLocalGroups(host, username)
            users.append(username)
            membership.append(group_member)
            status = win32net.NetUserGetInfo(host, username , 4)     
            #all_status.append(status)
            if int(hex(status['flags'])[-1]) == 1:
                account_one_status = 'enabled'
            elif int(hex(status['flags'])[-1]) == 3:
                account_one_status = 'disabled'
            account_status.update({username:account_one_status})
        Account_check_result.update({host:account_status})
    except:
        users.append('connection or access issue')
        membership.append('connection or access issue')
        Account_check_result.update({host:"connection or access issue"})
        #print('Issue with following IP address:   ' + str(host) +'\n')
        user_listt = [{'name':'test'}]
    for t in range(len(user_listt)):
        printer.append(user_listt[t]['name'])
    print('\n '.join(printer))

    return Account_check_result, users, membership

ip_scope = []


how_to_check = input('How to check hosts man/file?: ')
while how_to_check:
    if how_to_check == 'man':
        print('\n ok \n')
        break
    elif how_to_check == 'file':
        print('\n ok \n')
        break
    else:
       how_to_check = input('Wrong input! \n How to check hosts man/file?: ')


if how_to_check == 'man':
    host_list = input('Enter hostname or ip address or subnet: ')
    mask =input("Enter mask( n if no mask): ")
    if mask =='n':
        ip_scope.append(host_list)
    else:
        ip_scope = create_ip_list(host_list, mask)
elif how_to_check =='file':
    wb_read = load_workbook(filename='C:\\cisco_test\\IP_all.xlsx')
    ws_read = wb_read['Лист1']
    for i in range(5100):
        if ws_read.cell(row=i+1,column=1).value != None and isinstance(ws_read.cell(row=i+1,column=1).value, str):
            ip_scope.append(ws_read.cell(row=i+1,column=1).value)
    host_list = ("From excel last " + str(ip_scope[-1]))

result_list_users = open('C:\\cisco_test\\temp\\Result '+ 'net check ' + str(host_list) +'_local_user_checked.txt', mode ='a')
result_list_folders = open('C:\\cisco_test\\temp\\Result '+ 'net check '+ str(host_list) +'_shared_folder_checked.txt', mode ='a')
wb_write = load_workbook(filename='C:\\cisco_test\\Result '+ 'net check' +'.xlsx')


fill_cell_orange = PatternFill(start_color='f07233',
               end_color='f07233',
               fill_type='solid')
alignment_cell = Alignment(horizontal='left', vertical ='center', wrapText='True')

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))




if str(host_list) in wb_write.sheetnames:
    ws_write = wb_write[str(host_list)]
else:
    ws_write = wb_write.create_sheet(str(host_list))
    
ws_write.column_dimensions['A'].width = '21'
ws_write.column_dimensions['B'].width = '29'
ws_write.column_dimensions['C'].width = '32'
ws_write.column_dimensions['D'].width = '75'
ws_write.column_dimensions['E'].width = '70'
ws_write.column_dimensions['F'].width = '65'

ws_write.cell(row=1, column=1).value = 'IP адрес/Имя Хоста'
ws_write.cell(row=1, column=2).value = 'Расшареные сетевые папки'
ws_write.cell(row=1, column=3).value = ' Список локальных УЗ'
ws_write.cell(row=1, column=4).value = ' Список управления доступом'
ws_write.cell(row=1, column=5).value = ' Подпапки'
ws_write.cell(row=1, column=6).value = ' Группы локальных Юзеров'


for i in range(1,6):
    ws_write.cell(row=1, column=i).alignment=alignment_cell
    ws_write.cell(row=1, column=i).border = thin_border


for i in range(len(ip_scope)):
    if ws_write.cell(row=i+2,column=2).value == 'connection or access issue' or ws_write.cell(row=i+2,column=2).value == None:
        m = []
        k =[]
        ppp=[]
        host = ip_scope[i]
        Account_check_result, users, membership = check_pc_locusers(host)
        (result, permissions, turn) = host_folders(host)
        ws_write.cell(row=i+2, column=1).value = host
        ws_write.cell(row=i+2, column=1).alignment=alignment_cell
        if isinstance((result[host]), str):
            ws_write.cell(row=i+2, column=2).value = (result[host])
            ws_write.cell(row=i+2, column=2).alignment=alignment_cell
        else:
            ws_write.cell(row=i+2, column=2).value = ('\n'.join(result[host]))
            ws_write.cell(row=i+2, column=2).alignment=alignment_cell
        if isinstance((Account_check_result[host]), str):
            gh = Account_check_result[host]
            ws_write.cell(row=i+2, column=3).value = gh
            ws_write.cell(row=i+2, column=3).alignment=alignment_cell
        else:
            for key,value in Account_check_result[host].items():
                m.append(str(key) + '\t  ' +str(value))
            ws_write.cell(row=i+2, column=3).value = '; \n'.join(m)
            ws_write.cell(row=i+2, column=3).alignment=alignment_cell
       
        ws_write.cell(row=i+2, column=4).value = '\n'.join(permissions)
        ws_write.cell(row=i+2, column=4).alignment=alignment_cell
        if 'Все:(OI)(CI)(F)' in (''.join(permissions)) or 'Все:(OI)(CI)(IO)(GR,GE)' in (''.join(permissions)) or 'Все:(RX)' in (''.join(permissions)) or 'Все:(OI)(CI)(RX)' in (''.join(permissions)) or 'Все:(OI)(CI)(R)' in (''.join(permissions)) or 'Все:(OI)(CI)(M)' in (''.join(permissions)) or 'Все:(OI)(CI)(W)' in (''.join(permissions)) or 'Все:(CI)(RX)' in (''.join(permissions)):
            ws_write.cell(row=i+2, column=4).fill = fill_cell_orange
        
        ws_write.cell(row=i+2, column=5).value = '\n'.join(turn)
        ws_write.cell(row=i+2, column=5).alignment=alignment_cell
        
        if users[0] == 'connection or access issue':
            ws_write.cell(row=i+2, column=6).value = 'connection issue'
        else:
            for p in range(len(users)):
                ppp.append(str(users[p]) + ':   ' + '; '.join(membership[p]))
            ws_write.cell(row=i+2, column=6).value = '\n'.join(ppp)
            ws_write.cell(row=i+2, column=6).alignment=alignment_cell
    else:
        ws_write.cell(row=i+2, column=1).value = ws_write.cell(row=i+2,column=1).value
        ws_write.cell(row=i+2, column=2).value = ws_write.cell(row=i+2,column=2).value
        ws_write.cell(row=i+2, column=3).value = ws_write.cell(row=i+2,column=3).value
        ws_write.cell(row=i+2, column=4).value = ws_write.cell(row=i+2,column=4).value
        ws_write.cell(row=i+2, column=5).value = ws_write.cell(row=i+2,column=5).value
        ws_write.cell(row=i+2, column=6).value = ws_write.cell(row=i+2,column=6).value
        ws_write.cell(row=i+2, column=1).alignment=alignment_cell
        ws_write.cell(row=i+2, column=2).alignment=alignment_cell
        ws_write.cell(row=i+2, column=3).alignment=alignment_cell
        ws_write.cell(row=i+2, column=4).alignment=alignment_cell
        ws_write.cell(row=i+2, column=5).alignment=alignment_cell
        ws_write.cell(row=i+2, column=6).alignment=alignment_cell
    for k in range(1, 7):
        ws_write.cell(row=i+2, column=k).border = thin_border
    

wb_write.save(filename='C:\\cisco_test\\Result '+ 'net check.xlsx')
    





