from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side
import openpyxl

wb_active = load_workbook(filename='C:\\cisco_test\\cisco_users.xlsx')
ws_active = wb_active['0812']
wb_IT = load_workbook(filename='C:\\cisco_test\\net_dev_splunk.xlsx')
ws_IT = wb_IT['Лист1']
test_data_1 = []
test_data_2 = []
test_data_3 = []
test_data_4 = []
for i in range(1000):
    for m in range(1000):
        if ws_active.cell(row=m+1,column=1).value == ws_IT.cell(row=i+1, column=3).value:
            #ws_IT.cell(row=i+3, column=44).value = ws_active.cell(row=m+4,column=4).value
            #ws_IT.cell(row=i+3, column=45).value = ws_active.cell(row=m+4,column=5).value
            ws_active.cell(row=m+1,column=7).value =  ws_IT.cell(row=i+1, column=1).value
            test_data_1.append(ws_IT.cell(row=i+3,column=11).value)
            test_data_2.append(ws_IT.cell(row=i+3, column=2).value)
            test_data_3.append(ws_IT.cell(row=m+2,column=2).value)
            #print('Yes')
#ws_IT.cell(row=1, column=47).value = 
wb_active.save(filename='C:\\cisco_test\\cisco_users_with_names.xlsx')
