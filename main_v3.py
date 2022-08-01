import pynetbox
import requests
import csv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

from requests.packages.urllib3.exceptions import InsecureRequestWarning
session = requests.Session()
session.verify = False
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

nb = pynetbox.api(
    'https://netbox.infra.clouddc.ru/',
    token = '1578ba25aa0f7cacb5deeffdf77bf4df82e52fd6'
    )
nb.http_session = session
filename_csv_sibin = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\host_discovery sibin.csv'
filename_csv_infra = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\host_discovery_infra.csv'
file_audit = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\audit_vm.xlsx'
filename_siem = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\siem.csv'
filename_siem_actives = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\siem actives.csv'
filename_siem_src = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\siem sources.csv'
filename_tenable_sc = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\Assets vuln report.csv'
filename_vmm_report = r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\vmmreport.csv'
sheet_audit= "Main"
infra_audit_date = input('Введите дату последнего аудита ВМ в инфра формат (хх.хх.хх): ')
class asset_host_discovery:
    def __init__(self, ip, dns): #''', siem, nessus, avz, cyberark'''):
        self.ip = ip
        self.dns = dns
        self.tag = "Host Discovery"
        self.power_state = None
        self.service_model = None
        self.siem = None
        self.nessus = None
        self.avz = None
        self.cyberark = None
        self.dns_audit = None
        self.dns_netbox = None
        self.count = None
        self.IS = None
        self.src = None
        self.vuln_score = None
        self.total_score = None

class asset_nessus:
    def __init__(self, ip, power_state, dns): #''', siem, nessus, avz, cyberark'''):
        self.ip = ip
        self.tag = 'Netbox'
        self.dns = dns
        self.power_state = power_state
        self.service_model = None
        self.siem = None
        self.nessus = None
        self.avz = None
        self.cyberark = None


def open_siem_csv(filename_siem):
    ip_list =[]
    count_list = []
    with open(filename_siem, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter =';')
        for row in reader:
            if 'event_src.host' in row:
                for i in range(len(row)):
                    if row[i] =='event_src.host':
                        ip_addr = i
                    if row[i] =='COUNT':
                        count = i
                continue
            #print(row[name])
            ip_list.append(row[ip_addr].lower())
            count_list.append(row[count])
    return ip_list, count_list

def open_siem_actives_csv(filename_siem):
    ip_list =[]
    dns_siem_list = []
    with open(filename_siem, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter =';')
        for row in reader:
            if 'Host.@UpdateTime' in row:
                for i in range(len(row)):
                    if row[i] == r'host.IpAddress':
                        ip_addr = i
                    if row[i] == r'@Host':
                        host = i
                continue
            #print(row[name])
            ip_list.append(row[ip_addr])
            try:
                dns_name = row[host].split(' (')[0].lower()
            except:
                dns_name = row[host].lower()
            #print(dns_name)
            if dns_name == 'cdc-term-63.sibintek.ru':
                #print(dns_name)
                pass
            dns_siem_list.append(dns_name)
    return ip_list, dns_siem_list

def open_siem_src_csv(filename_siem_src):
    dns_list =[]
    count_list_src = []
    product_list = []
    with open(filename_siem_src, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter =';')
        for row in reader:
            if 'контроль отправки' in row:
                for i in range(len(row)):
                    if row[i] =='источник':
                        ip_addr = i
                    if row[i] =='количество событий':
                        count = i
                    if row[i] =='поставщик':
                        product = i
                continue
            #print(row[name])
            #print(row[ip_addr])
            try:
                dns_list.append(row[ip_addr].split(' (')[1].strip(')').lower())
            except:
                dns_list.append(row[ip_addr].lower())
            count_list_src.append(row[count])
            product_list.append(row[product])
    return dns_list, count_list_src, product_list
'''
reader = csv.reader(open("sibin_vuln.csv"))
reader1 = csv.reader(open("cloud_vuln.csv"))
f = open("combined_vuln.csv", "w")
writer = csv.writer(f)

for row in reader:
    writer.writerow(row)
for row in reader1:
    writer.writerow(row)
f.close()
'''
def open_nessus_vuln_csv(filename_tenable_sc):
    ip_list =[]
    score_list = []
    total_list = []
    with open(filename_tenable_sc, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter =',')
        for row in reader:
            if 'NetBIOS Name' in row:
                for i in range(len(row)):
                    if row[i] =='IP Address':
                        ip_addr = i
                    if row[i] =='Score':
                        score = i
                    if row[i] == 'Total':
                        total = i
                continue
            #print(row[name])
            #print(row[ip_addr])
            ip_list.append(row[ip_addr])
            score_list.append(row[score])
            total_list.append(row[total])
    return ip_list, score_list, total_list
def open_vmmreoprt_csv(filename_vmmreport):
    dns_list =[]
    ip_list = []
    with open(filename_vmmreport, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter = ',')
        for row in reader:
            if 'ipv4Addresses' in row:
                for i in range(len(row)):
                    if row[i] =='ipv4Addresses':
                        ip_addr = i
                    if row[i] =='Name':
                        dns = i
                continue
            #print(row[name])
            ip_list.append(row[ip_addr].lower())
            dns_name = row[dns].split('.')[0].lower()
            dns_list.append(dns_name)
    return ip_list, dns_list

ip_vmmreport, dns_vmmreport = open_vmmreoprt_csv(filename_vmm_report)
ip_sc_vuln, score_sc_vuln, total_sc_vuln = open_nessus_vuln_csv(filename_tenable_sc)
ip_siem, count_siem = open_siem_csv(filename_siem)
ip_siem_actives, dns_siem_actives = open_siem_actives_csv(filename_siem_actives)
dns_siem_src, dns_siem_count_src, product_siem_src = open_siem_src_csv(filename_siem_src)

def save_data_to_excel(*args, filename=r'U:\scripts\netbox plus nessus\main\result.xlsx', sheet_name='Main',
                       **kwargs):
    fill_cell_orange = PatternFill(start_color='f07233',
                                   end_color='f07233',
                                   fill_type='solid')
    alignment_cell = Alignment(horizontal='left', vertical='center', wrapText='True')

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(sheet_name)
    for counter, k in enumerate(kwargs):
        # print(counter)
        # print(k)
        # print(type(k))
        ws_write.cell(row=1, column=counter + 1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter + 2, column=count_data + 1).value = data[counter]
            ws_write.cell(row=counter + 2, column=count_data + 1).alignment = alignment_cell
            ws_write.cell(row=counter + 2, column=count_data + 1).border = thin_border
    wb_write.save(filename)


def open_host_discovery(filename):
    ip_dict_v ={}
    with open(filename, newline='', encoding='utf-8') as File:
        reader = csv.reader(File, delimiter =',')
        for row in reader:
            if 'IP Address' in row:
                for i in range(len(row)):
                    if row[i] =='IP Address':
                        ip_addr = i
                    if row[i] == 'DNS Name':
                        dns = i
                continue
            #print(row[name])
            #print(row[dns] + " type: " + str(type(row[dns])))
            ip_dict_v.update({row[ip_addr]:asset_host_discovery(row[ip_addr], row[dns])})
            #asset_hd = asset_host_discovery(row)
    return ip_dict_v
ip_dict_1 = open_host_discovery(filename_csv_sibin)
ip_dict_2 = open_host_discovery(filename_csv_infra)
ip_dict ={}
ip_dict.update(ip_dict_1)
ip_dict.update(ip_dict_2)


# Сбор данных с Нетбокс
#arr_4 = nb.ipam.ip_addresses.all()
#
#
#
arr_2 = nb.virtualization.virtual_machines.all()
ip_netbox = []
name_netbox = []
status_netbox = []
ip_dict_netbox = {}
for i in range(len(arr_2)):
    name_netbox.append(arr_2[i]['name'])
    if arr_2[i]['primary_ip4'] is None:
        ip_var = None
    else:
        ip_temp = arr_2[i]['primary_ip4']['address']
        ip_var = ip_temp.split('/')[0]
    ip_netbox.append(ip_var)
    status_netbox.append(arr_2[i]['status']['value'])
    ip_dict_netbox.update({ip_var:asset_nessus(ip_var, arr_2[i]['status']['value'], arr_2[i]['name'])})


class asset_audit:
    def __init__(self, ip, power_state, dns, service_model, avz): #''', siem, nessus, avz, cyberark'''):
        self.ip = ip
        self.power_state = power_state
        self.dns = dns
        self.service_model = service_model
        self.siem = None
        self.nessus = None
        self.avz =avz
        self.cyberark = None

def get_data_from_excel(filename, sheet_name, column_1, column_2, column_3, column_4, column_5, range_row=5000):
    data_1 = []
    data_2 = []
    data_3 = []
    data_4 = []
    data_5 = []
    wb_excel = load_workbook(filename)
    ws_excel = wb_excel[sheet_name]
    for i in range(range_row):
        if ws_excel.cell(row=i+1, column=1).value is not None:
            data_1.append(ws_excel.cell(row=i+1, column=column_1).value)
            data_2.append(ws_excel.cell(row=i+1, column=column_2).value)
            data_3.append(ws_excel.cell(row=i+1, column=column_3).value)
            data_4.append(ws_excel.cell(row=i+1, column=column_4).value)
            data_5.append(ws_excel.cell(row=i+1, column=column_5).value)
    wb_excel.save(filename)
    return data_1, data_2, data_3, data_4, data_5

audit_list_ip, audit_list_vm_name, audit_list_pow_st, audit_list_kasp_status, audit_list_model = get_data_from_excel(file_audit, sheet_audit, 9, 1, 8, 4, 10)
audit_list_ip_infra, audit_list_vm_name_infra, audit_list_pow_st_infra, audit_list_kasp_status_infra, audit_list_model_infra = get_data_from_excel(r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\CloudDC&INFRA ALL SRV and Antivirus.xlsx', infra_audit_date, 10, 1, 3, 4, 11)
audit_list_ip_extra, audit_list_vm_name_extra, audit_list_pow_st_extra, audit_list_kasp_status_extra, audit_list_model_extra = get_data_from_excel(file_audit, 'extra', 9, 1, 8, 4, 10)
audit_list_ip, audit_list_vm_name, audit_list_pow_st, audit_list_kasp_status, audit_list_model = audit_list_ip + audit_list_ip_infra,\
                                                                                                 audit_list_vm_name + audit_list_vm_name_infra, audit_list_pow_st + audit_list_pow_st_infra,\
                                                                                                 audit_list_kasp_status+audit_list_kasp_status_infra, audit_list_model+ ['SaaS_clouddc' for i in range(len(audit_list_model_infra))]
audit_list_ip, audit_list_vm_name, audit_list_pow_st, audit_list_kasp_status, audit_list_model = audit_list_ip + audit_list_ip_extra,\
                                                                                                 audit_list_vm_name + audit_list_vm_name_extra, audit_list_pow_st + audit_list_pow_st_extra,\
                                                                                                 audit_list_kasp_status+audit_list_kasp_status_extra, audit_list_model+ ['Management' for i in range(len(audit_list_model_extra))]


system_names, description_ip_list, test1, test2, test3 = get_data_from_excel(r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\input files\description.xlsx', 'Main', 1, 3, 4, 5, 6)
ip_dict_auditvm ={}
'''
for i in range(len(audit_list_vm_name)):
    ip_dict_auditvm.update({audit_list_vm_name:asset_audit(audit_list_ip, audit_list_pow_st, audit_list_vm_name, audit_list_model, audit_list_kasp_status)})
'''
# Сравнение всех данных

for i in ip_dict.keys():
    for m in range(len(audit_list_vm_name)):
        if ip_dict[i].ip == audit_list_ip[m] or ip_dict[i].dns.split('.')[0] == audit_list_vm_name[m].lower() or ip_dict[i].dns == audit_list_vm_name[m].lower():
            ip_dict[i].power_state = audit_list_pow_st[m]
            ip_dict[i].avz = audit_list_kasp_status[m]
            ip_dict[i].service_model = audit_list_model[m]
            ip_dict[i].dns_audit = audit_list_vm_name[m].split('.')[0].lower()
            
for i in ip_dict.keys():
    for m in range(len(ip_vmmreport)):
        if ip_dict[i].ip == ip_vmmreport[m] and (ip_dict[i].dns_audit is None or ip_dict[i].dns_audit == ""):
            ip_dict[i].dns_audit = dns_vmmreport[m]
            
for i in ip_dict.keys():
    for m in range(len(audit_list_vm_name)):
        if ip_dict[i].dns_audit == audit_list_vm_name[m].lower() and (ip_dict[i].power_state is None or ip_dict[i].power_state ==""):
            ip_dict[i].power_state = audit_list_pow_st[m]
            ip_dict[i].avz = audit_list_kasp_status[m]
            ip_dict[i].service_model = audit_list_model[m]
            ip_dict[i].dns_audit = audit_list_vm_name[m].split('.')[0].lower()

for i in ip_dict.keys():
    for m in range(len(name_netbox)):
        if ip_dict[i].ip == ip_netbox[m]:
            ip_dict[i].dns_netbox = name_netbox[m]
            #print('i')
        if ip_dict[i].ip == ip_netbox[m] and (ip_dict[i].power_state is None or ip_dict[i].power_state ==""):
            ip_dict[i].power_state = status_netbox[m]
            #print('i' + 's_part')

#Сравнение кол-ва событий внизу эталонный вариант
'''
for i in ip_dict.keys():
    for m in range(len(ip_siem)):
        if ip_dict[i].ip == ip_siem[m] or ip_dict[i].dns == ip_siem[m] or ip_dict[i].dns_audit == ip_siem[m].split('.')[0]:            
            ip_dict[i].count = count_siem[m]
'''
for i in ip_dict.keys():
    count_final = 'src'
    for m in range(len(ip_siem)):
        if ip_dict[i].ip == ip_siem[m]:
            count_final += ' + ' + str(count_siem[m])
            continue
        if ip_dict[i].dns == ip_siem[m] or ip_dict[i].dns.split('.')[0] == ip_siem[m].split('.')[0]:
            count_final += ' + ' + str(count_siem[m])
            continue
        if ip_dict[i].dns_audit == ip_siem[m].split('.')[0]:
            count_final += ' + ' + str(count_siem[m])
            continue
        if ip_dict[i].dns_netbox == ip_siem[m]:
            count_final += ' + ' + str(count_siem[m])
            continue

        if ip_dict[i].dns_netbox is not None and ip_dict[i].dns_netbox.split('.')[0] == ip_siem[m].split('.')[0]:
            count_final += ' + ' + str(count_siem[m])
            continue

    count_result = count_final if count_final!= 'src' else 0
    ip_dict[i].count = count_result

for i in ip_dict.keys():
    for m in range(len(ip_siem_actives)):
        if ip_dict[i].ip == ip_siem_actives[m] or ip_dict[i].dns == dns_siem_actives[m] or ip_dict[i].dns_audit == dns_siem_actives[m].split('.')[0].lower() or ip_dict[i].dns_netbox == dns_siem_actives[m].lower():
            ip_dict[i].siem = 'SIEM'

for i in ip_dict.keys():
    for m in range(len(description_ip_list)):
        if ip_dict[i].ip == description_ip_list[m]:
            ip_dict[i].IS = system_names[m]

for i in ip_dict.keys():
    count_product = ''
    for m in range(len(dns_siem_src)):
        if (ip_dict[i].ip == dns_siem_src[m] or ip_dict[i].dns==dns_siem_src[m] or ip_dict[i].dns_audit== dns_siem_src[m].split('.')[0] or ip_dict[i].dns_netbox== dns_siem_src[m])\
                and dns_siem_count_src[m] != 0 and ('snegirsoft' not in ip_dict[i].dns or 'snegirsoft' not in ip_dict[i].dns_netbox):
            count_product += str(product_siem_src[m]) + ':' + str(dns_siem_count_src[m]) + ';\n'
            continue
    count_product_result = count_product if count_product != '' else 0
    ip_dict[i].src = count_product_result
for i in ip_dict.keys():
    for m in range(len(ip_sc_vuln)):
        if ip_dict[i].ip == ip_sc_vuln[m]:
            ip_dict[i].vuln_score = 'Score:' + score_sc_vuln[m]
            ip_dict[i].total_score = 'Total:' + total_sc_vuln[m]

print(len(ip_dict.keys()))

#save_data_to_excel(ip_dict.ip)
def save_data_to_cell(ws_name, row, col, var):
    ws_name.cell(row=row, column=col).value = var
def save_final_data(filename= r'G:\Департамент информационной безопасности\14_ОЭСЗИ\1_Отчеты\Актуализация активов\Report\final.xlsx', sheet_name = 'Main'):
    try:
        wb_excel = load_workbook(filename)
    except FileNotFoundError:
        wb_excel = Workbook()
    ws_excel = wb_excel.create_sheet(sheet_name)
    col_names = ['Наименование ИС', 'IP адрес', 'DNS при сканировании', 'DNS ИМЯ', 'Имя в Netbox','Тег','Сервисная Модель','Статус','Касперский', 'СЦСАС', 'Событий за три дня', 'События', 'Оценка уязвимости', 'Кол-во уязвимостей']
    for i, k in enumerate(ip_dict):
        save_data_to_cell(ws_excel, i+2, 1, ip_dict[k].IS)
        save_data_to_cell(ws_excel, i+2, 2, ip_dict[k].ip)
        save_data_to_cell(ws_excel, i + 2, 3, ip_dict[k].dns)
        save_data_to_cell(ws_excel, i + 2, 4, ip_dict[k].dns_audit)
        save_data_to_cell(ws_excel, i + 2, 5, ip_dict[k].dns_netbox)
        save_data_to_cell(ws_excel, i + 2, 6, ip_dict[k].tag)
        save_data_to_cell(ws_excel, i + 2, 7, ip_dict[k].service_model)
        save_data_to_cell(ws_excel, i + 2, 8, ip_dict[k].power_state)
        save_data_to_cell(ws_excel, i + 2, 9, ip_dict[k].avz)
        save_data_to_cell(ws_excel, i + 2, 10, ip_dict[k].siem)
        save_data_to_cell(ws_excel, i + 2, 11, ip_dict[k].count)
        save_data_to_cell(ws_excel, i + 2, 12, ip_dict[k].src)
        save_data_to_cell(ws_excel, i + 2, 13, ip_dict[k].vuln_score)
        save_data_to_cell(ws_excel, i + 2, 14, ip_dict[k].total_score)
    for i in range(1,15):
        save_data_to_cell(ws_excel, 1, i, col_names[i-1])
    wb_excel.save(filename)
save_final_data()
