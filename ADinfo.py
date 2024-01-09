from typing import List

from ldap3 import Server, Connection, ALL, NTLM, ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES, ObjectDef, AttrDef, Reader, \
    Writer, Entry, Attribute, OperationalAttribute, SUBTREE, core
from ldap3.core import exceptions
import ldap3
import os
import csv
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color, colors, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side

import re
import datetime

pswd = 'dfjidfindifdfdf'
sAMAccountName = []

now = datetime.datetime.now()
if now.month < 10:
        month_string = '0' + str(now.month)
else:
        month_string = str(now.month)

if now.day < 10:
        day_string = '0' + str(now.day)
else:
        day_string = str(now.day)
date_string = day_string + month_string

server = Server('ldap://dc.loc', port=389, use_ssl=True, get_info=ALL)
conn = Connection(server, user="DC.loc\avkuzmin", password=pswd, authentication=NTLM)
conn.bind()
conn.start_tls()



def username_list(*args, filename=None, filetype='splunk', **kwargs):
    sAMAccountName = []
    usernames = []
    csvdata = []
    if filename is None:
        for arg in args:
            usernames.append(arg)
    elif isinstance(filename, str):
        if filetype == 'splunk':
            file_obj = open(filename, encoding='utf-8')
            reader = csv.reader(file_obj, dialect='excel', delimiter=',')
            for row in reader:
                csvdata.append(row)
            csvdata_index = csvdata[0].index('user')
            for item in csvdata:
                usernames.append(item[csvdata_index])
            usernames.pop(0)
        elif filetype == 'excel':
            wb = load_workbook(filename)
            ws = wb["Лист1"]
            for i in range(7000):
                usernames.append(ws.cell(row=i + 1, column=1).value)
    return usernames


def surname_list(*args, filename=None, filetype='splunk', **kwargs):
    sAMAccountName = []
    surnames = []
    csvdata = []
    if filename is None:
        for arg in args:
            surnames.append(arg)
    elif isinstance(filename, str):
        if filetype == 'splunk':
            file_obj = open(filename, encoding='utf-8')
            reader = csv.reader(file_obj, dialect='excel', delimiter=',')
            for row in reader:
                csvdata.append(row)
            csvdata_index = csvdata[0].index('fio')
            for item in csvdata:
                surnames.append(item[csvdata_index])
            surnames.pop(0)
        elif filetype == 'excel':
            wb = load_workbook(filename)
            ws = wb["Лист1"]
            for i in range(7000):
                surnames.append(ws.cell(row=i + 1, column=1).value)
    return surnames


def getinfo_by_username(username):
    global conn
    sAMAccountName = '(sAMAccountName=' + username + ')'
    try:
        #conn.search('DC=DC, DC =LOC', sAMAccountName, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
        conn.extend.standard.paged_search('DC=DC, DC =LOC', sAMAccountName, search_scope=SUBTREE,
                                          attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        all_data = conn.entries[0]
    except IndexError:
        all_data = 'No such user with username: ' + username + ' in AD'
    return all_data


def getinfo_by_surname(surname):
    global conn
    name = '(name=' + surname + ')'
    try:
        #conn.search('DC=DC, DC =LOC', name, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
        conn.extend.standard.paged_search('DC=DC, DC =LOC', name, search_scope=SUBTREE,
                                          attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        all_data = conn.entries[0]
    except IndexError:
        all_data = 'No such user with username: ' + surname + ' in AD'
    return all_data


def get_all_workers(filename):
    wb_read = load_workbook(filename)
    ws_read = wb_read['Текущие сотрудники на дату']
    all_users = []
    for i in range(1, 8000):
        if ws_read.cell(row=i + 2, column=2).value != None and isinstance(ws_read.cell(row=i + 2, column=2).value, str):
            all_users.append(ws_read.cell(row=i + 2, column=2).value)
    return all_users


def get_status_by_username(username):
    global conn
    name = '(sAMAccountName=' + username + ')'
    try:
        #conn.search('DC=DC, DC =LOC', name, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
        conn.extend.standard.paged_search('DC=DC, DC =LOC', name, search_scope=SUBTREE,
                                          attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        if conn.entries[0].userAccountControl.value == 514 or conn.entries[0].userAccountControl.value == 546 or \
                conn.entries[0].userAccountControl.value == 66050 or conn.entries[0].userAccountControl.value == 66082:
            status = 'Locked'
        else:
            status = 'Unlocked'
    except IndexError:
        status = 'No such user'
    return status


def get_status_by_surname(surname):
    global conn
    name = '(name=' + surname + ')'
    try:
        #conn.search('DC=DC, DC =LOC', name, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
        conn.extend.standard.paged_search('DC=DC, DC =LOC', name, search_scope=SUBTREE,
                                          attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        if conn.entries[0].userAccountControl.value == 514 or conn.entries[0].userAccountControl.value == 546 or \
                conn.entries[0].userAccountControl.value == 66050 or conn.entries[0].userAccountControl.value == 66082:
            status = 'Locked'
        else:
            status = 'Unlocked'
    except IndexError:
        status = 'No such user'
    return status


def get_users_from_group(groupname):
    global conn
    # name = '(sAMAccountName=' + username + ')'
    groupname = '(cn=' + str(groupname) + ')'
    # search_base = str(groupname) + ',OU=Groups,DC=SVO,DC=AIR,DC=LOC'
    users_in_group = []
    try:
        #conn.search('DC=DC, DC =LOC', groupname, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
        conn.extend.standard.paged_search('DC=SVO,DC=AIR,DC=LOC', groupname, search_scope=SUBTREE, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        for i in range(len(conn.entries[0].member.value)):
            result = re.search(r'^CN=(.+?),', conn.entries[0].member.value[i])
            member = result.group(1)
            users_in_group.append(member)
    except ldap3.core.exceptions.LDAPCursorError:
        conn.extend.standard.paged_search('DC=DC, DC =LOC', groupname, search_scope=SUBTREE, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        for i in range(len(conn.entries[1].member.value)):
            result = re.search(r'^CN=(.+?),', conn.entries[1].member.value[i])
            member = result.group(1)
            users_in_group.append(member)
    except:
        users_in_group.append('Issue')
    return users_in_group

#def get_users_from_specified_containers(*container=['OU=07_01,OU=SVO_new,DC=SVO,DC=AIR,DC=LOC']
def get_users_from_specified_containers(container):
    global conn
    username =[]
    surname = []
    for ou in container:
        try:
            #conn.search(ou, '(objectclass=user)', attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
            conn.extend.standard.paged_search(ou, '(objectclass=user)', search_scope=SUBTREE,
                                              attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
            for i in range(len(conn.entries)):
                username.append(conn.entries[i].sAMAccountName.value)
                surname.append(conn.entries[i].name.value)
        except:
            username.append('Issue')
            surname.append('Issue')
    return username, surname
def get_users_with_specified_attribute(name, value):
    global conn
    username = []
    surname = []
    search_filter ='(&(' + str(name) + '=' + str(value) + ')' + '(objectclass=user))'
    try:
        #conn.search('DC=DC, DC =LOC', search_filter, paged_size=4000, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES])
        conn.extend.standard.paged_search('DC=DC, DC =LOC', search_filter, search_scope=SUBTREE, attributes=[ALL_ATTRIBUTES, ALL_OPERATIONAL_ATTRIBUTES], generator=False)
        for i in range(len(conn.entries)):
            username.append(conn.entries[i].sAMAccountName.value)
            surname.append(conn.entries[i].name.value)
    except:
        username.append('Issue')
        surname.append('Issue')
    if len(username) == 0:
        username.append('No such users')
    if len(surname) == 0:
        surname.append('No such users')
    return username, surname
def save_data_to_excel(*args, filename='c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\results\\Проверка12.xlsx', **kwargs):
    try:
        wb_write = load_workbook(filename)
    except FileNotFoundError:
        wb_write = Workbook()
    ws_write = wb_write.create_sheet(str(date_string))
    for counter, k in enumerate(kwargs):
        print(counter)
        print(k)
        print(type(k))
        ws_write.cell(row=1, column=counter+1).value = kwargs[k]
    for count_data, data in enumerate(args):
        for counter, info in enumerate(data):
            ws_write.cell(row=counter+2, column=count_data+1).value = data[counter]

    wb_write.save(filename)
    return wb_write


def excel_ccs(func):
    def wrapper(*args, filename='c:\\Users\\avkuzmin\\Documents\\scripts\\AD\\results\\Проверка12.xlsx', **kwargs):
        wb_work = func(*args,**kwargs)
        ws_write = wb_work[str(date_string)]
        thin_border = Border(left=Side(style='thin'),
                               right=Side(style='thin'),
                               top=Side(style='thin'),
                               bottom=Side(style='thin'))
        alignment_cell = Alignment(horizontal='left', vertical='center', wrapText='True')
        fill_cell_orange = PatternFill(start_color='f07233',
                                           end_color='f07233',
                                           fill_type='solid')
        for count_data, data in enumerate(args):
            for counter, info in enumerate(data):
                ws_write.cell(row=counter + 2, column=count_data + 1).border = thin_border
                ws_write.cell(row=counter + 2, column=count_data + 1).alignment = alignment_cell
        wb_work.save(filename)
        return wb_work
    return wrapper
